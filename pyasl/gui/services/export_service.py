"""
gui/services/export_service.py
--------------------------------
Export service for results — CSV, JSON, Excel, PDF, chart images.

Pure Python — no Qt dependency for data transforms.
Chart image export requires matplotlib (already a dependency).
"""
from __future__ import annotations

import csv
import io
import json
import logging
import os
from typing import Any, Dict, List, Optional

from pyasl.gui.models.result_data import DatasetEntry, ProcessedResult
from pyasl.gui.models.result_ui_state import ResultUIState

logger = logging.getLogger(__name__)


class ResultExporter:
    """Export processed results in various formats."""

    # ------------------------------------------------------------------
    # CSV
    # ------------------------------------------------------------------

    def export_csv(
        self,
        entries: List[DatasetEntry],
        path: str,
        include_header: bool = True,
    ) -> None:
        """Export entries as CSV."""
        if not entries:
            with open(path, "w", newline="", encoding="utf-8") as f:
                f.write("No data\n")
            return

        fieldnames = [
            "name", "file_path", "shape", "dtype", "ndim",
            "min_val", "max_val", "mean_val", "std_val",
            "size", "description",
        ]

        with open(path, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            if include_header:
                writer.writeheader()
            for entry in entries:
                row = {
                    "name": entry.name,
                    "file_path": entry.file_path,
                    "shape": str(entry.shape) if entry.shape else "",
                    "dtype": entry.dtype,
                    "ndim": entry.ndim,
                    "min_val": entry.min_val if entry.min_val is not None else "",
                    "max_val": entry.max_val if entry.max_val is not None else "",
                    "mean_val": entry.mean_val if entry.mean_val is not None else "",
                    "std_val": entry.std_val if entry.std_val is not None else "",
                    "size": entry.size,
                    "description": entry.description,
                }
                writer.writerow(row)

        logger.info("CSV exported to %s (%d entries)", path, len(entries))

    # ------------------------------------------------------------------
    # JSON
    # ------------------------------------------------------------------

    def export_json(
        self,
        result: ProcessedResult,
        path: str,
        filtered_entries: Optional[List[DatasetEntry]] = None,
    ) -> None:
        """Export result as JSON."""
        data = result.to_dict()
        if filtered_entries is not None:
            data["entries"] = [e.to_dict() for e in filtered_entries]
            data["_filtered"] = True
            data["_total_entries"] = result.entry_count

        with open(path, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2, ensure_ascii=False)

        logger.info("JSON exported to %s", path)

    # ------------------------------------------------------------------
    # Excel
    # ------------------------------------------------------------------

    def export_excel(
        self,
        entries: List[DatasetEntry],
        result: ProcessedResult,
        path: str,
    ) -> None:
        """
        Export entries as Excel (.xlsx).

        Uses openpyxl if available, falls back to CSV.
        """
        try:
            import openpyxl
        except ImportError:
            logger.warning("openpyxl not installed — falling back to CSV.")
            csv_path = path.replace(".xlsx", ".csv")
            self.export_csv(entries, csv_path)
            return

        wb = openpyxl.Workbook()

        # Sheet 1: Data
        ws = wb.active
        ws.title = "Data"

        headers = [
            "Name", "Shape", "Type", "Dimensions",
            "Min", "Max", "Mean", "Std Dev",
            "Elements", "Description", "File Path",
        ]
        ws.append(headers)

        # Style header
        from openpyxl.styles import Font, PatternFill
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="830085", end_color="830085", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill

        for entry in entries:
            ws.append([
                entry.name,
                str(entry.shape) if entry.shape else "",
                entry.dtype,
                entry.ndim,
                entry.min_val,
                entry.max_val,
                entry.mean_val,
                entry.std_val,
                entry.size,
                entry.description,
                entry.file_path,
            ])

        # Sheet 2: Summary
        ws2 = wb.create_sheet("Summary")
        ws2.append(["Metric", "Value"])
        ws2.append(["Result Name", result.name])
        ws2.append(["Result ID", result.result_id])
        ws2.append(["Source", result.source_reference])
        ws2.append(["Created", result.created_at])
        ws2.append(["Total Records", result.entry_count])
        ws2.append(["ASL Version", result.processing_metadata.asl_version])
        if result.processing_metadata.processing_duration is not None:
            ws2.append([
                "Processing Duration",
                f"{result.processing_metadata.processing_duration:.2f}s",
            ])

        # Auto-width columns
        for ws_active in [ws, ws2]:
            for col in ws_active.columns:
                max_len = 0
                col_letter = col[0].column_letter
                for cell in col:
                    try:
                        cell_len = len(str(cell.value or ""))
                        max_len = max(max_len, cell_len)
                    except Exception:
                        pass
                ws_active.column_dimensions[col_letter].width = min(max_len + 2, 50)

        wb.save(path)
        logger.info("Excel exported to %s (%d entries)", path, len(entries))

    # ------------------------------------------------------------------
    # PDF Report
    # ------------------------------------------------------------------

    def export_pdf_report(
        self,
        result: ProcessedResult,
        entries: List[DatasetEntry],
        path: str,
        summary_metrics: Optional[Dict[str, Any]] = None,
        chart_image_path: Optional[str] = None,
    ) -> None:
        """
        Generate a PDF report using matplotlib's PDF backend.

        Contains: header, metadata, summary metrics, optional chart,
        data table summary.
        """
        import matplotlib
        matplotlib.use("Agg")
        from matplotlib.backends.backend_pdf import PdfPages
        import matplotlib.pyplot as plt

        with PdfPages(path) as pdf:
            # Page 1: Title and Summary
            fig, ax = plt.subplots(figsize=(8.5, 11))
            ax.axis("off")

            # Title
            ax.text(0.5, 0.95, result.name, fontsize=20, fontweight="bold",
                    ha="center", va="top", transform=ax.transAxes)
            ax.text(0.5, 0.91, f"Result ID: {result.result_id}",
                    fontsize=9, ha="center", va="top", color="#666",
                    transform=ax.transAxes)
            ax.text(0.5, 0.88, f"Source: {result.source_reference}",
                    fontsize=9, ha="center", va="top", color="#666",
                    transform=ax.transAxes)

            # Metadata
            y = 0.82
            meta = result.processing_metadata
            info_lines = [
                f"ASL Library: {meta.asl_library} v{meta.asl_version}",
                f"Pipeline: {meta.pipeline_name} ({meta.pipeline_type})",
                f"Created: {result.created_at}",
                f"Total Records: {result.entry_count}",
            ]
            if meta.processing_duration is not None:
                info_lines.append(f"Processing Duration: {meta.processing_duration:.2f}s")

            for line in info_lines:
                ax.text(0.1, y, line, fontsize=10, va="top",
                        transform=ax.transAxes)
                y -= 0.03

            # Summary metrics
            if summary_metrics:
                y -= 0.03
                ax.text(0.1, y, "Summary Metrics", fontsize=12,
                        fontweight="bold", va="top", transform=ax.transAxes)
                y -= 0.03
                for key, val in summary_metrics.items():
                    ax.text(0.12, y, f"{key}: {val}", fontsize=10,
                            va="top", transform=ax.transAxes)
                    y -= 0.025

            # Data table summary
            y -= 0.04
            ax.text(0.1, y, "Dataset Entries", fontsize=12,
                    fontweight="bold", va="top", transform=ax.transAxes)
            y -= 0.03

            for entry in entries[:30]:  # Limit to 30 entries in PDF
                line = f"• {entry.name} — {entry.dtype} {entry.shape}"
                if entry.mean_val is not None:
                    line += f" — mean={entry.mean_val:.4f}"
                ax.text(0.12, y, line, fontsize=8, va="top",
                        transform=ax.transAxes)
                y -= 0.02
                if y < 0.05:
                    break

            if len(entries) > 30:
                ax.text(0.12, y, f"... and {len(entries) - 30} more entries",
                        fontsize=8, color="#888", va="top",
                        transform=ax.transAxes)

            pdf.savefig(fig)
            plt.close(fig)

            # Page 2: Chart (if available)
            if chart_image_path and os.path.isfile(chart_image_path):
                fig2, ax2 = plt.subplots(figsize=(8.5, 11))
                ax2.axis("off")
                img = plt.imread(chart_image_path)
                ax2.imshow(img)
                ax2.set_title("Visualization", fontsize=14, pad=20)
                pdf.savefig(fig2)
                plt.close(fig2)

        logger.info("PDF report exported to %s", path)

    # ------------------------------------------------------------------
    # Chart image export
    # ------------------------------------------------------------------

    def export_chart_image(
        self,
        figure: Any,  # matplotlib Figure
        path: str,
        fmt: str = "png",
        dpi: int = 150,
    ) -> None:
        """Save a matplotlib figure to an image file."""
        figure.savefig(path, format=fmt, dpi=dpi, bbox_inches="tight",
                       facecolor=figure.get_facecolor(), edgecolor="none")
        logger.info("Chart exported to %s (format=%s)", path, fmt)

    # ------------------------------------------------------------------
    # Convenience: current view vs full result
    # ------------------------------------------------------------------

    def export_current_view(
        self,
        result: ProcessedResult,
        filtered_entries: List[DatasetEntry],
        path: str,
        fmt: str = "csv",
    ) -> None:
        """Export what the user currently sees (filtered)."""
        if fmt == "csv":
            self.export_csv(filtered_entries, path)
        elif fmt == "json":
            self.export_json(result, path, filtered_entries)
        elif fmt == "xlsx":
            self.export_excel(filtered_entries, result, path)
        else:
            self.export_csv(filtered_entries, path)

    def export_full_result(
        self,
        result: ProcessedResult,
        path: str,
        fmt: str = "csv",
    ) -> None:
        """Export the complete canonical dataset."""
        if fmt == "csv":
            self.export_csv(result.entries, path)
        elif fmt == "json":
            self.export_json(result, path)
        elif fmt == "xlsx":
            self.export_excel(result.entries, result, path)
        else:
            self.export_csv(result.entries, path)
